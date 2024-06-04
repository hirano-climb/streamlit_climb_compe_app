import streamlit as st
import sqlite3
import openpyxl as op
from io import BytesIO

def create_excel():
    
    wb = op.Workbook()
    ws = wb.worksheets[0]
    
    ws.cell(2, 1).value = 'ニックネーム'
    ws.cell(2, 2).value = '出場カテゴリー'
    ws.cell(2, 3).value = '出場時間帯'
    ws.cell(2, 4).value = '合計得点'

    for i in range(1, 37):
        ws.cell(2, 4+i).value = f"課題{i}"
        
        if f"課題{i}" == '課題1' or f"課題{i}" == '課題5':
            ws.cell(1, 4+i).value = 20
        elif f"課題{i}" == '課題3' or f"課題{i}" == '課題6':
            ws.cell(1, 4+i).value = 100
        else:
            ws.cell(1, 4+i).value = 5


    DATABASE = './data/profile_result.db'
    con = sqlite3.connect(DATABASE)
    c = con.cursor()
    select_value_column = 'SELECT * FROM profile_result'
    c.execute(select_value_column)

    for j, row in enumerate(c):       
        ws.cell(j+3, 1).value = row[0] 
        ws.cell(j+3, 2).value = row[1] 
        ws.cell(j+3, 3).value = row[2]       
        total_list = []
        num = row[3].split(',')
        
        for k in range(1, 37): 
            ws.cell(j+3, 4+k).value = 0
            
            for l in num:        
                if l == str(k):
                    ws.cell(j+3, 4+k).value = 1
                
            total_list.append(ws.cell(1, 4+k).value*ws.cell(j+3, 4+k).value) 
        
        ws.cell(j+3, 4).value = sum(total_list)          
        
    c.close()
    con.close()
    
    wb.save('./data/result.xlsx')


def download_excel():
    wb = op.load_workbook('./data/result.xlsx')
    output = BytesIO()
    wb.save(output)
    xl_btn= st.download_button(
        "Download",
        output.getvalue(),
        "リザルト.xlsx"
        )
    
    if xl_btn:
        st.success('入力データの出力が完了しました')
    

if __name__ == '__main__':
    create_excel()
    download_excel()

